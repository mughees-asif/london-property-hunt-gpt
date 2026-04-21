"""XLSX tracker persistence for scored listings from @pipeline."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from property_hunt.models import Listing, ListingType, Priority, canonicalize_url


HEADERS = [
    "Title",
    "Platform",
    "URL",
    "Area",
    "Postcode",
    "Price (pcm)",
    "Bills Included",
    "Available From",
    "Furnished",
    "Bed Count",
    "Flatmates",
    "Contact",
    "Notes",
    "Status",
    "Priority",
    "Found On",
]

ROOM_SHEET = "Listings"
STUDIO_SHEET = "Studios & 1-Beds"

PRIORITY_FILLS = {
    Priority.HIGH: "E2EFDA",
    Priority.MEDIUM: "FFFFC7",
    Priority.LOW: "FCE4D6",
}


@dataclass(frozen=True)
class TrackerResult:
    """Summary of tracker writes consumed by @email.render and run reports."""

    added: list[Listing]
    duplicates: list[Listing]


def init_tracker(path: Path) -> None:
    """Create a fresh tracker workbook at the requested path."""

    wb = _new_workbook()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def append_new_listings(path: Path, listings: list[Listing]) -> TrackerResult:
    """Append unseen trackable listings and return added/duplicate splits."""

    wb = _load_or_create(path)
    seen_urls = existing_urls(wb)
    added: list[Listing] = []
    duplicates: list[Listing] = []

    for listing in listings:
        if not listing.is_trackable:
            continue
        canonical = listing.canonical_url
        if canonical in seen_urls:
            duplicates.append(listing)
            continue
        ws = wb[_sheet_for(listing)]
        row_index = ws.max_row + 1
        ws.append(listing.to_tracker_row())
        _format_listing_row(ws, row_index, listing)
        seen_urls.add(canonical)
        added.append(listing)

    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return TrackerResult(added=added, duplicates=duplicates)


def existing_urls(wb: object) -> set[str]:
    """Load canonical URLs already present in both tracker sheets."""

    urls: set[str] = set()
    for sheet_name in (ROOM_SHEET, STUDIO_SHEET):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            value = row[0]
            if value:
                urls.add(canonicalize_url(str(value)))
    return urls


def _load_or_create(path: Path) -> object:
    """Load an existing workbook or initialise a compatible one."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read/write tracker spreadsheets") from exc

    if path.exists():
        wb = load_workbook(path)
        _ensure_workbook_shape(wb)
        return wb
    return _new_workbook()


def _new_workbook() -> object:
    """Build an empty workbook with all expected sheets and formatting."""

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read/write tracker spreadsheets") from exc

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for sheet_name in (ROOM_SHEET, STUDIO_SHEET):
        wb.create_sheet(sheet_name)
    _ensure_workbook_shape(wb)
    return wb


def _ensure_workbook_shape(wb: object) -> None:
    """Create missing sheets and apply header-level formatting."""

    for sheet_name in (ROOM_SHEET, STUDIO_SHEET):
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            ws.append(HEADERS)
        _format_header(ws)
        _set_column_widths(ws)
        ws.freeze_panes = "A2"


def _format_header(ws: object) -> None:
    """Apply tracker header values and styles."""

    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    for index, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=index)
        cell.value = header
        cell.fill = fill
        cell.font = font


def _set_column_widths(ws: object) -> None:
    """Set readable default column widths for manual review."""

    widths = {
        "A": 45,
        "B": 12,
        "C": 50,
        "D": 18,
        "E": 10,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 10,
        "K": 24,
        "L": 20,
        "M": 44,
        "N": 12,
        "O": 10,
        "P": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _format_listing_row(ws: object, row_index: int, listing: Listing) -> None:
    """Apply priority fill and hyperlink style to one appended row."""

    from openpyxl.styles import PatternFill

    if listing.priority in PRIORITY_FILLS:
        fill = PatternFill("solid", fgColor=PRIORITY_FILLS[listing.priority])
        for cell in ws[row_index]:
            cell.fill = fill
    url_cell = ws.cell(row=row_index, column=3)
    url_cell.hyperlink = listing.url
    url_cell.style = "Hyperlink"


def _sheet_for(listing: Listing) -> str:
    """Route room and whole-unit listings to their tracker sheets."""

    return ROOM_SHEET if listing.listing_type == ListingType.ROOM else STUDIO_SHEET
